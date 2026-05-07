#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.10"
# dependencies = [
#   "openpyxl",
# ]
# ///

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Sequence

from ruleset_compiler.ir import (
    LiteralPattern,
    MatchContext,
    Rule,
    RuleSource,
    RulesetIR,
)


DEFAULT_RECORD_SIZE = 64
DEFAULT_WINDOW_SIZE = 8
DEFAULT_CONTEXT_ID = 1


@dataclass(frozen=True, slots=True)
class IPSWorkbookSchema:
    sheet_name: str = "IPS_패턴"
    code_column: str = "패턴코드"
    name_column: str = "공격명"
    protocol_column: str = "프로토콜"
    port_column: str = "포트"
    pattern_type_column: str = "패턴유형"
    pattern_column: str = "탐지문자열"
    offset_column: str = "옵셋값"
    offset_cmp_column: str = "옵셋비교"
    case_sensitive_column: str = "대소문자구별"
    priority_column: str = "위험도"

    def resolve_columns(self, header: Sequence[Any]) -> dict[str, int]:
        header_map = {
            str(name).strip(): index
            for index, name in enumerate(header)
            if name is not None
        }
        column_names = {
            "code": self.code_column,
            "name": self.name_column,
            "protocol": self.protocol_column,
            "port": self.port_column,
            "pattern_type": self.pattern_type_column,
            "pattern": self.pattern_column,
            "offset": self.offset_column,
            "offset_cmp": self.offset_cmp_column,
            "case_sensitive": self.case_sensitive_column,
            "priority": self.priority_column,
        }
        missing = [name for name in column_names.values() if name not in header_map]
        if missing:
            raise ValueError(
                "Workbook schema mismatch: missing columns "
                + ", ".join(sorted(missing))
            )
        return {
            field: header_map[column_name]
            for field, column_name in column_names.items()
        }


def decode_pattern(raw: str, pattern_type: str) -> bytes:
    pattern_type_norm = pattern_type.strip().upper()
    if pattern_type_norm == "BIN":
        compact = "".join(raw.replace("%", " ").split())
        return bytes.fromhex(compact)
    return raw.encode("latin-1", errors="replace")


def parse_ips_workbook(
    workbook_path: Path,
    *,
    schema: IPSWorkbookSchema = IPSWorkbookSchema(),
) -> RulesetIR:
    try:
        import openpyxl
    except ImportError as error:
        raise RuntimeError(
            "openpyxl is required to parse IPS workbook files. "
            "Install it or run this parser via uv with openpyxl available."
        ) from error

    workbook = openpyxl.load_workbook(str(workbook_path), read_only=True)
    try:
        if schema.sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet {schema.sheet_name!r} not found in {workbook_path}; "
                f"available sheets: {', '.join(workbook.sheetnames)}"
            )
        worksheet = workbook[schema.sheet_name]
        return parse_ips_rows(
            worksheet.iter_rows(values_only=True),
            source_uri=str(workbook_path),
            schema=schema,
        )
    finally:
        workbook.close()


def parse_ips_rows(
    rows: Iterable[Sequence[Any]],
    *,
    source_uri: str = "<rows>",
    schema: IPSWorkbookSchema = IPSWorkbookSchema(),
) -> RulesetIR:
    row_iter = iter(rows)
    header = next(row_iter, None)
    if header is None:
        return RulesetIR()

    columns = schema.resolve_columns(header)
    ir = RulesetIR()
    source_id = ir.next_source_id()
    ir.sources.append(
        RuleSource(
            source_id=source_id,
            source_type="ips_workbook",
            uri=source_uri,
            native_engine="custom_ips_workbook",
            metadata={"sheet_name": schema.sheet_name},
        )
    )
    ir.contexts.append(
        MatchContext(
            context_id=DEFAULT_CONTEXT_ID,
            protocol="any",
            buffer_kind="payload",
            normalization="raw",
            direction="either",
            stream_scope="packet",
        )
    )

    for source_line, row in enumerate(row_iter, start=2):
        raw_pattern = _cell(row, columns["pattern"])
        if raw_pattern in (None, ""):
            continue

        pattern_type = _text(_cell(row, columns["pattern_type"])) or "TEXT"
        try:
            pattern_bytes = decode_pattern(_text(raw_pattern), pattern_type)
        except ValueError:
            continue
        if not pattern_bytes:
            continue

        native_id = _text(_cell(row, columns["code"])) or f"row:{source_line}"
        protocol = _text(_cell(row, columns["protocol"])) or "any"
        port = _text(_cell(row, columns["port"]))
        severity = _text(_cell(row, columns["priority"])) or None
        message = _text(_cell(row, columns["name"])) or None
        offset = _parse_int(_cell(row, columns["offset"]))
        offset_cmp = _text(_cell(row, columns["offset_cmp"])) or None
        case_sensitive = _text(_cell(row, columns["case_sensitive"]))

        rule_uid = ir.next_rule_uid()
        ir.rules.append(
            Rule(
                rule_uid=rule_uid,
                source_id=source_id,
                native_id=native_id,
                native_engine="custom_ips_workbook",
                severity=severity,
                message=message,
                source_line=source_line,
                metadata={
                    "protocol": protocol,
                    "port": port,
                    "offset_cmp": offset_cmp,
                },
            )
        )

        ir.patterns.append(
            LiteralPattern(
                pattern_uid=ir.next_pattern_uid(),
                rule_uid=rule_uid,
                context_id=DEFAULT_CONTEXT_ID,
                data=pattern_bytes,
                nocase=_is_nocase(case_sensitive),
                offset=offset,
                pattern_type=pattern_type.strip().upper(),
                source_line=source_line,
                metadata={
                    "source_pattern_text": _text(raw_pattern),
                    "protocol": protocol,
                    "port": port,
                },
            )
        )

    return ir


def _cell(row: Sequence[Any], index: int) -> Any:
    return row[index] if index < len(row) else None


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_int(value: Any) -> int | None:
    text = _text(value)
    if not text:
        return None
    try:
        return int(text, 0)
    except ValueError:
        return None


def _is_nocase(case_sensitive: str) -> bool:
    value = case_sensitive.strip().lower()
    if not value:
        return False
    case_sensitive_values = {"y", "yes", "true", "1", "o", "yes", "case", "sensitive"}
    case_insensitive_values = {"n", "no", "false", "0", "x", "nocase", "insensitive"}
    if value in case_sensitive_values:
        return False
    if value in case_insensitive_values:
        return True
    if value in {"예", "유", "구별", "구분"}:
        return False
    if value in {"아니오", "무", "미구별", "구별안함", "구분안함"}:
        return True
    return False
