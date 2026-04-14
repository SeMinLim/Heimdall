#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.10"
# dependencies = [
#   "openpyxl",
# ]
# ///
"""Export IPS workbook patterns to the HPAT binary format.

Usage:
    uv run export_hpat.py --ips-xlsx path/to/workbook.xlsx
    uv run export_hpat.py --ips-xlsx path/to/workbook.xlsx --output data/real_patterns.bin

The HPAT format is the interchange format between this Python exporter
and the C++ pattern selector reference implementation.  See the C++ header
reference/include/heimdall/benchmark_data.hpp for the binary layout.
"""

from __future__ import annotations

import argparse
import struct
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Sequence


DEFAULT_RECORD_SIZE = 64
DEFAULT_WINDOW_SIZE = 8

REAL_PATTERN_FILE_MAGIC = b"HPAT"
REAL_PATTERN_FILE_VERSION = 1
_HEADER_STRUCT = struct.Struct(
    "<4sHHHI"
)  # magic + version + record_size + window_size + num_records
_ENTRY_HEADER_STRUCT = struct.Struct("<H")  # pattern_len per entry


@dataclass(frozen=True, slots=True)
class IPSRule:
    code: Any
    name: str
    protocol: Any
    port: Any
    pattern_type: str
    offset: Any
    offset_cmp: Any
    case_sensitive: str
    priority: Any
    raw_bytes: bytes


@dataclass(slots=True)
class PreparedPatternDataset:
    records: list[bytes]
    pattern_lens: list[int]
    skipped_records: int = 0


@dataclass(frozen=True, slots=True)
class IPSWorkbookSchema:
    sheet_name: str = "IPS_패턴"
    code_column: str = "패턴코드"
    name_column: str = "공격명"
    protocol_column: str = "프로토콜"
    port_column: str = "포트"
    pattern_type_column: str = "패턴유형"
    pattern_column: str = "탐지문자열"
    offset_column: str = "옵셋값"
    offset_cmp_column: str = "옵셋비교"
    case_sensitive_column: str = "대소문자구별"
    priority_column: str = "위험도"

    def resolve_columns(self, header: Sequence[Any]) -> dict[str, int]:
        header_map = {
            str(name).strip(): index
            for index, name in enumerate(header)
            if name is not None
        }
        column_names = {
            "code": self.code_column,
            "name": self.name_column,
            "protocol": self.protocol_column,
            "port": self.port_column,
            "pattern_type": self.pattern_type_column,
            "pattern": self.pattern_column,
            "offset": self.offset_column,
            "offset_cmp": self.offset_cmp_column,
            "case_sensitive": self.case_sensitive_column,
            "priority": self.priority_column,
        }
        missing = [name for name in column_names.values() if name not in header_map]
        if missing:
            raise ValueError(
                "Workbook schema mismatch: missing columns "
                + ", ".join(sorted(missing))
            )
        return {
            field: header_map[column_name]
            for field, column_name in column_names.items()
        }


def decode_pattern(raw: str, pattern_type: str) -> bytes:
    """Decode a pattern string from the IPS workbook format.

    BIN type uses ``%XX`` hex notation; everything else is latin-1 text.
    """
    if pattern_type == "BIN":
        return bytes.fromhex(raw.replace("%", ""))
    return raw.encode("latin-1", errors="replace")


def load_ips_patterns(workbook_path: Path) -> list[IPSRule]:
    import openpyxl

    schema = IPSWorkbookSchema()
    workbook = openpyxl.load_workbook(str(workbook_path), read_only=True)
    try:
        if schema.sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet {schema.sheet_name!r} not found in {workbook_path}; "
                f"available sheets: {', '.join(workbook.sheetnames)}"
            )
        worksheet = workbook[schema.sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            return []

        col = schema.resolve_columns(header)
        rules: list[IPSRule] = []

        for row in rows:
            raw_pattern = row[col["pattern"]]
            if not raw_pattern:
                continue

            pattern_type = str(row[col["pattern_type"]] or "")
            try:
                pattern_bytes = decode_pattern(str(raw_pattern), pattern_type)
            except (ValueError, UnicodeDecodeError):
                continue

            rules.append(
                IPSRule(
                    code=row[col["code"]],
                    name=str(row[col["name"]] or ""),
                    protocol=row[col["protocol"]],
                    port=row[col["port"]],
                    pattern_type=pattern_type,
                    offset=row[col["offset"]],
                    offset_cmp=row[col["offset_cmp"]],
                    case_sensitive=str(row[col["case_sensitive"]] or "").strip(),
                    priority=row[col["priority"]],
                    raw_bytes=pattern_bytes,
                )
            )

        return rules
    finally:
        workbook.close()


def build_real_record_dataset(
    rules: Sequence[IPSRule],
    record_size: int = DEFAULT_RECORD_SIZE,
    window_size: int = DEFAULT_WINDOW_SIZE,
) -> PreparedPatternDataset:
    records: list[bytes] = []
    pattern_lens: list[int] = []
    skipped = 0

    for rule in rules:
        raw = rule.raw_bytes
        if len(raw) > record_size:
            records.append(raw[:record_size])
            pattern_lens.append(record_size)
        elif len(raw) >= window_size:
            records.append(raw.ljust(record_size, b"\x00"))
            pattern_lens.append(len(raw))
        else:
            skipped += 1

    return PreparedPatternDataset(
        records=records,
        pattern_lens=pattern_lens,
        skipped_records=skipped,
    )


def write_real_patterns_binary(
    dataset: PreparedPatternDataset,
    out_path: Path,
    record_size: int = DEFAULT_RECORD_SIZE,
    window_size: int = DEFAULT_WINDOW_SIZE,
) -> int:
    if len(dataset.records) != len(dataset.pattern_lens):
        raise ValueError("records and pattern_lens must have the same length")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    payload = bytearray(
        _HEADER_STRUCT.pack(
            REAL_PATTERN_FILE_MAGIC,
            REAL_PATTERN_FILE_VERSION,
            record_size,
            window_size,
            len(dataset.records),
        )
    )

    for record, pattern_len in zip(dataset.records, dataset.pattern_lens):
        if len(record) != record_size:
            raise ValueError(
                f"record length must be {record_size} bytes, got {len(record)}"
            )
        if pattern_len < 0 or pattern_len > record_size:
            raise ValueError(
                f"pattern_len must be within 0..{record_size}, got {pattern_len}"
            )

        payload += _ENTRY_HEADER_STRUCT.pack(pattern_len)
        payload += record

    out_path.write_bytes(payload)
    return len(dataset.records)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export IPS workbook patterns to HPAT binary format.",
    )
    parser.add_argument(
        "--ips-xlsx",
        type=Path,
        required=True,
        help="Path to the IPS workbook (.xlsx).",
    )
    parser.add_argument(
        "--output",
        "-o",
        type=Path,
        default=Path("data/real_patterns.bin"),
        help="Output path for the HPAT binary (default: data/real_patterns.bin).",
    )
    parser.add_argument(
        "--record-size",
        type=int,
        default=DEFAULT_RECORD_SIZE,
        help=f"Fixed record slot size in bytes (default: {DEFAULT_RECORD_SIZE}).",
    )
    parser.add_argument(
        "--window-size",
        type=int,
        default=DEFAULT_WINDOW_SIZE,
        help=f"Fingerprint window size in bytes (default: {DEFAULT_WINDOW_SIZE}).",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    print(f"Loading IPS patterns from: {args.ips_xlsx}")
    rules = load_ips_patterns(args.ips_xlsx.expanduser())
    print(f"  Loaded {len(rules)} rules")

    lengths = [len(rule.raw_bytes) for rule in rules]
    if lengths:
        print(
            f"  Pattern length: min={min(lengths)}, "
            f"avg={sum(lengths) / len(lengths):.1f}, max={max(lengths)}"
        )

    dataset = build_real_record_dataset(
        rules,
        record_size=args.record_size,
        window_size=args.window_size,
    )

    if dataset.skipped_records:
        print(f"  Skipped (< {args.window_size}B): {dataset.skipped_records}")

    records_written = write_real_patterns_binary(
        dataset,
        out_path=args.output,
        record_size=args.record_size,
        window_size=args.window_size,
    )

    print(f"\nExported {records_written} records to {args.output}")
    print(f"  File size: {args.output.stat().st_size:,} bytes")


if __name__ == "__main__":
    main()
